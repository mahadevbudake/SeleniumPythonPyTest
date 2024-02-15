import pytest
import openpyxl

class ExcelReadWrite:
    def OpenandReadExcel(self):
        # to load the workbook with its path
        bk = openpyxl.load_workbook("C:\\Users\\mahadev_budake\\Downloads\\Screenshots\\TestData.xlsx")
        # to identify active worksheet
        s = bk.active
        print(s.max_row)
        c = s.cell(row=3, column=2).value
        print(c)
        return c
    def OpenandWriteExcel(self):
        # to load the workbook with its path
        wb = openpyxl.load_workbook("C:\\Users\\mahadev_budake\\Downloads\\Screenshots\\TestData.xlsx")
        # to identify active worksheet
        #s = bk.active
        ws = wb.worksheets[0]
        ws.cell(row=4, column=4).value = 2
        wb.save("C:\\Users\\mahadev_budake\\Downloads\\Screenshots\\TestData.xlsx")